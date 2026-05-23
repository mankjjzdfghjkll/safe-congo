import openpyxl
from openpyxl import load_workbook

wb = load_workbook('data/raw/drc-2022_sem40.xlsx', read_only=True, data_only=True)
print('Feuilles:', wb.sheetnames)

ws = wb.active


print('\n--- 30 premiers en-têtes ---')
row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
cols = [(i+1, v) for i, v in enumerate(row1) if v is not None]
for i, v in cols[:30]:
    print(f'  Col {i}: {v}')
print(f'  ... total colonnes non nulles: {len(cols)}')

print('\n--- 5 premières lignes (20 premières colonnes) ---')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, max_col=20, values_only=True)):
    print(f'  Ligne {i+2}:', list(row))

wb.close()
